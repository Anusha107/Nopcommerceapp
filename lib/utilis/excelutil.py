import openpyxl
def get_Rowcount(filename,sheetname):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook[sheetname]
    return sheet.max_row

def get_Columncount(filename,sheetname):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook[sheetname]
    return sheet.max_column

def read_data(filename,sheetname,rowNumber,coloumnNumber):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook[sheetname]
    return sheet.cell(row=rowNumber,column=coloumnNumber).value

def write_data(filename,sheetname,rowNumber,coloumnNumber,data):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook[sheetname]
    sheet.cell(row=rowNumber,column=coloumnNumber).value =data 
    workbook.save()
